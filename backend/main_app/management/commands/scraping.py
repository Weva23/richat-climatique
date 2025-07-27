import re
import requests
from bs4 import BeautifulSoup
from pathlib import Path
import html
import ftfy
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from django.core.management.base import BaseCommand
from urllib.parse import urljoin, urlparse, parse_qs
from datetime import datetime
import time
import socket
import logging
import traceback
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import (
    NoSuchElementException, 
    TimeoutException,
    ElementClickInterceptedException,
    WebDriverException,
    StaleElementReferenceException
)
from webdriver_manager.chrome import ChromeDriverManager
from typing import List, Dict, Optional

# Configuration du logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class Command(BaseCommand):
    help = 'Scraper intégré GEF et GCF pour les projets de Mauritanie - FORMAT COMPATIBLE'

    def __init__(self):
        super().__init__()
        self.total_projects = 0
        self.scraped_projects = 0
        self.driver = None
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        })

    def add_arguments(self, parser):
        parser.add_argument(
            '--source',
            type=str,
            choices=['gef', 'gcf', 'both'],
            help='Source à scraper: gef, gcf ou both',
            default='both'
        )
        parser.add_argument(
            '--max-pages',
            type=int,
            help='Nombre maximum de pages à scraper (pour GEF uniquement)',
            default=50
        )
        parser.add_argument(
            '--headless',
            action='store_true',
            help='Exécuter en mode headless (sans interface)',
            default=False
        )
        parser.add_argument(
            '--max-details',
            type=int,
            help='Nombre maximum de projets à enrichir en détail (pour GCF)',
            default=None
        )

    def handle(self, *args, **options):
        source = options.get('source', 'both')
        max_pages = options.get('max_pages', 50)
        headless = options.get('headless', False)
        max_details = options.get('max_details', None)
        
        self.stdout.write("🚀 SCRAPER INTÉGRÉ GEF-GCF MAURITANIE")
        self.stdout.write("=" * 75)
        self.stdout.write("🎯 SOURCES DISPONIBLES:")
        self.stdout.write("   - GEF: Global Environment Facility")
        self.stdout.write("   - GCF: Green Climate Fund")
        self.stdout.write(f"📊 SOURCE SÉLECTIONNÉE: {source.upper()}")
        self.stdout.write("💾 FORMAT: Compatible entre GEF et GCF")
        self.stdout.write("=" * 75)
        
        try:
            all_projects = []
            
            if source in ['gef', 'both']:
                self.stdout.write("\n🌍 === SCRAPING GEF MAURITANIE ===")
                self.setup_driver(headless)
                gef_projects = self.scrape_gef_mauritania_projects(max_pages)
                all_projects.extend(gef_projects)
                self.cleanup_driver()
            
            if source in ['gcf', 'both']:
                self.stdout.write("\n🌱 === SCRAPING GCF MAURITANIE ===")
                gcf_projects = self.scrape_gcf_mauritania_projects(max_details)
                all_projects.extend(gcf_projects)
            
            # Sauvegarde séparée par source
            if all_projects:
                self.save_projects_by_source(all_projects, source)
                
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"❌ Erreur générale: {e}"))
            traceback.print_exc()
        finally:
            self.cleanup_driver()
        
        self.stdout.write(self.style.SUCCESS(
            f"\n🎉 SCRAPING TERMINÉ!\n"
            f"📊 Total projets scrapés: {len(all_projects) if 'all_projects' in locals() else 0}\n"
            f"📁 Fichiers de sortie:\n"
            f"   - GEF_Mauritanie_Projects.xlsx (projets GEF)\n"
            f"   - GCF_Mauritanie_Projects.xlsx (projets GCF)\n"
            f"✅ Sauvegarde séparée par source!"
        ))

    def clean_text(self, text):
        """Nettoyer le texte des espaces excessifs et caractères spéciaux"""
        if not text:
            return ""
        text = str(text)
        text = html.unescape(text)
        text = ftfy.fix_text(text)
        return " ".join(text.split()).strip()

    # =============================
    # MÉTHODES GEF (avec Selenium)
    # =============================

    def setup_driver(self, headless=False):
        """Configuration du driver Chrome pour GEF"""
        try:
            options = Options()
            options.add_argument("--start-maximized")
            options.add_argument("--disable-infobars")
            options.add_argument("--disable-extensions")
            options.add_argument("--disable-notifications")
            options.add_argument("--disable-popup-blocking")
            options.add_argument("--no-sandbox")
            options.add_argument("--disable-dev-shm-usage")
            options.add_argument("--disable-gpu")
            options.add_argument("--window-size=1920,1080")
            options.add_argument("--disable-blink-features=AutomationControlled")
            options.add_experimental_option("excludeSwitches", ["enable-automation"])
            options.add_experimental_option('useAutomationExtension', False)
            options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
            
            if headless:
                options.add_argument("--headless")
            
            service = Service(ChromeDriverManager().install())
            self.driver = webdriver.Chrome(service=service, options=options)
            self.driver.implicitly_wait(15)
            
            self.driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
            
            logger.info("✅ Driver Chrome configuré pour GEF")
            
        except Exception as e:
            logger.error(f"❌ Erreur configuration driver: {e}")
            raise

    def cleanup_driver(self):
        """Nettoyer proprement le driver"""
        if self.driver:
            try:
                self.driver.quit()
                logger.info("🔒 Driver fermé")
            except Exception as e:
                logger.warning(f"⚠ Erreur fermeture driver: {e}")

    def scrape_gef_mauritania_projects(self, max_pages):
        """Scraper principal pour les projets GEF Mauritanie"""
        base_url = "https://www.thegef.org/projects-operations/database?f%5B0%5D=project_country_national%3A105"
        
        try:
            self.stdout.write(f"\n🌐 Navigation vers: {base_url}")
            self.driver.get(base_url)
            
            WebDriverWait(self.driver, 30).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "table.views-table, .view-content"))
            )
            
            # Gérer les cookies
            try:
                cookie_btn = WebDriverWait(self.driver, 5).until(
                    EC.element_to_be_clickable((By.XPATH, "//button[contains(., 'Accept') or contains(., 'Accepter')]"))
                )
                self.driver.execute_script("arguments[0].click();", cookie_btn)
                time.sleep(2)
            except:
                pass
            
            page_num = 1
            projects = []
            
            while page_num <= max_pages:
                self.stdout.write(f"\n📄 TRAITEMENT PAGE GEF {page_num}")
                
                page_projects = self.extract_gef_projects_from_page()
                
                if not page_projects:
                    self.stdout.write("❌ Aucun projet GEF trouvé sur cette page")
                    break
                
                for project_basic in page_projects:
                    project_enriched = self.enrich_gef_project_details(project_basic)
                    projects.append(project_enriched)
                    
                    if len(projects) <= 5 or len(projects) % 5 == 0:
                        self.stdout.write(f"✅ Projet GEF #{len(projects)}: {project_enriched['Titre'][:50]}...")
                
                self.stdout.write(f"📊 Page GEF {page_num}: {len(page_projects)} projets extraits")
                
                if not self.navigate_to_next_page():
                    self.stdout.write("📄 Dernière page GEF atteinte")
                    break
                
                page_num += 1
                time.sleep(2)
            
            self.stdout.write(f"✅ GEF terminé: {len(projects)} projets extraits")
            return projects
            
        except Exception as e:
            self.stdout.write(f"❌ Erreur lors du scraping GEF: {e}")
            traceback.print_exc()
            return []

    def extract_gef_projects_from_page(self):
        """Extraire tous les projets GEF de la page actuelle"""
        projects = []
        
        try:
            WebDriverWait(self.driver, 15).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "table.views-table tbody tr, .view-content"))
            )
            
            soup = BeautifulSoup(self.driver.page_source, 'html.parser')
            table = soup.find('table', class_='views-table')
            if not table:
                return projects
            
            rows = table.find('tbody').find_all('tr') if table.find('tbody') else []
            
            for row in rows:
                try:
                    project_data = self.extract_gef_project_data_from_row(row)
                    if project_data:
                        projects.append(project_data)
                except Exception as e:
                    self.stdout.write(f"⚠ Erreur extraction ligne GEF: {e}")
                    continue
            
            return projects
            
        except Exception as e:
            self.stdout.write(f"❌ Erreur extraction projets page GEF: {e}")
            return projects

    def extract_gef_project_data_from_row(self, row):
        """Extraire les données d'un projet GEF depuis une ligne du tableau"""
        try:
            cells = row.find_all('td')
            if len(cells) < 9:
                return None
            
            title_cell = cells[0]
            title_link = title_cell.find('a')
            titre = self.clean_text(title_link.get_text()) if title_link else self.clean_text(title_cell.get_text())
            
            lien_projet = ""
            if title_link and title_link.get('href'):
                lien_projet = urljoin("https://www.thegef.org", title_link['href'])
            
            type_projet = self.clean_text(cells[4].get_text())
            organisation = self.clean_text(cells[5].get_text())
            cofinancement_total = self.clean_text(cells[7].get_text())
            
            if not titre or len(titre) < 10:
                return None
            
            return {
                'Titre': titre,
                'Type': type_projet,
                'Document': '',
                'nom_site': 'thegef.org',
                'Organisation': organisation,
                'Lien': lien_projet,
                'Description': '',
                'Cofinancement Total': cofinancement_total,
                'source': 'GEF',
                'gef_project_id': self.clean_text(cells[1].get_text()),
                'pays': self.clean_text(cells[2].get_text()),
                'domaines_focaux': self.clean_text(cells[3].get_text()),
                'statut': self.clean_text(cells[8].get_text())
            }
            
        except Exception as e:
            logger.error(f"Erreur extraction données projet GEF: {e}")
            return None

    def enrich_gef_project_details(self, project_basic):
        """Enrichir les détails du projet GEF depuis sa page dédiée"""
        enriched = project_basic.copy()
        
        if not project_basic.get('Lien'):
            return enriched
        
        try:
            self.driver.execute_script("window.open('');")
            self.driver.switch_to.window(self.driver.window_handles[-1])
            
            self.driver.get(project_basic['Lien'])
            
            WebDriverWait(self.driver, 15).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, ".field, .project-details"))
            )
            
            soup = BeautifulSoup(self.driver.page_source, 'html.parser')
            self.extract_gef_project_page_details(soup, enriched)
            
            self.driver.close()
            self.driver.switch_to.window(self.driver.window_handles[0])
            
            time.sleep(1)
            
        except Exception as e:
            self.stdout.write(f"⚠ Erreur enrichissement GEF: {e}")
            try:
                if len(self.driver.window_handles) > 1:
                    self.driver.close()
                    self.driver.switch_to.window(self.driver.window_handles[0])
            except:
                pass
        
        return enriched

    def extract_gef_project_page_details(self, soup, project):
        """Extraire les détails depuis la page projet GEF"""
        try:
            # Description
            description_selectors = [
                '.field--name-body .field__item',
                '.field--name-field-summary .field__item',
                '.project-summary',
                'meta[name="description"]'
            ]
            
            for selector in description_selectors:
                desc_elem = soup.select_one(selector)
                if desc_elem:
                    if desc_elem.name == 'meta':
                        project['Description'] = desc_elem.get('content', '')
                    else:
                        project['Description'] = self.clean_text(desc_elem.get_text())
                    break
            
            if not project['Description']:
                gef_id = project.get('gef_project_id', '')
                pays = project.get('pays', '')
                domaines = project.get('domaines_focaux', '')
                statut = project.get('statut', '')
                project['Description'] = f"Projet GEF {gef_id} en {pays}. Domaines focaux: {domaines}. Statut: {statut}."
            
            # Documents
            doc_section = soup.find('div', class_='field--name-field-document-url')
            if doc_section:
                doc_items = doc_section.find_all('div', class_='field__item')
                doc_urls = []
                
                for item in doc_items:
                    link_elem = item.find('a', href=True)
                    if link_elem:
                        doc_urls.append(link_elem['href'])
                
                if doc_urls:
                    project['Document'] = ' | '.join(doc_urls)
            
            # Autres enrichissements...
            financial_section = soup.find('div', class_='project-financials')
            if financial_section:
                cofinancing_elem = financial_section.find('div', class_='views-field-field-co-financing-total')
                if cofinancing_elem:
                    amount_elem = cofinancing_elem.find('div', class_='field-content')
                    if amount_elem:
                        amount = self.clean_text(amount_elem.get_text())
                        if amount and amount != project['Cofinancement Total']:
                            project['Cofinancement Total'] = f"USD {amount}"
            
        except Exception as e:
            logger.error(f"Erreur extraction détails page projet GEF: {e}")

    def navigate_to_next_page(self):
        """Naviguer vers la page suivante GEF"""
        try:
            time.sleep(2)
            
            next_selectors = [
                "//li[@class='page-item']/a[contains(@title, 'Go to next page') or contains(@rel, 'next')]",
                "//a[contains(@title, 'Go to next page')]",
                "//a[contains(@rel, 'next')]",
                "//li[contains(@class, 'page-item')]/a[contains(text(), '››')]",
                "//a[contains(@aria-label, 'Next')]"
            ]
            
            for selector in next_selectors:
                try:
                    next_link = self.driver.find_element(By.XPATH, selector)
                    if next_link.is_displayed() and next_link.is_enabled():
                        parent_li = next_link.find_element(By.XPATH, "./..")
                        if "disabled" not in parent_li.get_attribute("class").lower():
                            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", next_link)
                            time.sleep(1)
                            self.driver.execute_script("arguments[0].click();", next_link)
                            
                            WebDriverWait(self.driver, 15).until(
                                EC.presence_of_element_located((By.CSS_SELECTOR, "table.views-table tbody tr"))
                            )
                            
                            return True
                except:
                    continue
            
            return False
            
        except Exception as e:
            self.stdout.write(f"❌ Erreur navigation page suivante GEF: {e}")
            return False

    # =============================
    # MÉTHODES GCF (avec Requests)
    # =============================

    def scrape_gcf_mauritania_projects(self, max_details=None):
        """Scraper principal pour les projets GCF Mauritanie"""
        base_url = "https://www.greenclimate.fund"
        mauritania_url = "https://www.greenclimate.fund/countries/mauritania"
        
        try:
            self.stdout.write(f"\n🌐 Navigation vers: {mauritania_url}")
            
            # Récupération de tous les projets depuis les pages de liste
            projects = self.get_all_gcf_pages_projects(mauritania_url, base_url)
            self.stdout.write(f"Total des projets GCF trouvés: {len(projects)}")
            
            if not projects:
                self.stdout.write("❌ Aucun projet GCF trouvé")
                return []

            # Enrichissement avec les détails
            if max_details:
                projects = projects[:max_details]
                
            enriched_projects = self.enrich_gcf_projects_details(projects, base_url)
            
            self.stdout.write(f"✅ GCF terminé: {len(enriched_projects)} projets extraits")
            return enriched_projects
            
        except Exception as e:
            self.stdout.write(f"❌ Erreur lors du scraping GCF: {e}")
            traceback.print_exc()
            return []

    def get_page_content(self, url: str) -> Optional[BeautifulSoup]:
        """Récupère le contenu HTML d'une page GCF"""
        try:
            response = self.session.get(url, timeout=30)
            response.raise_for_status()
            return BeautifulSoup(response.content, 'html.parser')
        except requests.RequestException as e:
            logger.error(f"Erreur lors de la récupération de {url}: {e}")
            return None

    def get_all_gcf_pages_projects(self, mauritania_url, base_url):
        """Récupère les projets GCF de toutes les pages avec pagination"""
        all_projects = []
        page = 0
        
        while True:
            if page == 0:
                url = mauritania_url
            else:
                url = f"{mauritania_url}?page={page}"
            
            self.stdout.write(f"📄 TRAITEMENT PAGE GCF {page + 1}: {url}")
            soup = self.get_page_content(url)
            
            if not soup:
                break

            page_projects = self.extract_gcf_projects_from_table(soup, base_url)
            
            if not page_projects:
                self.stdout.write("❌ Aucun projet GCF trouvé sur cette page")
                break
            
            all_projects.extend(page_projects)
            self.stdout.write(f"📊 Page GCF {page + 1}: {len(page_projects)} projets trouvés")
            
            # Vérification pagination
            pagination = soup.find('ul', class_='pager')
            if pagination:
                next_link = pagination.find('a', title=re.compile(r'Go to next page|Aller à la page suivante'))
                if not next_link:
                    self.stdout.write("📄 Dernière page GCF atteinte")
                    break
            else:
                self.stdout.write("📄 Pas de pagination GCF trouvée")
                break
            
            page += 1
            time.sleep(1)

        return all_projects

    def extract_gcf_projects_from_table(self, soup: BeautifulSoup, base_url: str):
        """Extrait les projets GCF depuis le tableau de la page principale"""
        projects = []
        
        table = soup.find('table', class_='table')
        if not table:
            table = soup.find('table', class_=['views-table', 'sticky-enabled'])
            if not table:
                table = soup.find('table')
        
        if not table:
            logger.warning("❌ Tableau des projets GCF non trouvé")
            return projects

        tbody = table.find('tbody')
        if tbody:
            rows = tbody.find_all('tr')
        else:
            rows = table.find_all('tr')[1:]
        
        for i, row in enumerate(rows):
            try:
                cells = row.find_all('td')
                if len(cells) < 3:
                    continue

                title_cell = cells[0]
                link_elem = title_cell.find('a')
                if not link_elem:
                    continue

                project_url = urljoin(base_url, link_elem.get('href', ''))
                titre_base = link_elem.get_text(strip=True)

                document_type = cells[1].get_text(strip=True) if len(cells) > 1 else ''

                cover_date = ''
                if len(cells) > 2:
                    date_span = cells[2].find('span', class_='date-display-single')
                    if date_span:
                        cover_date = date_span.get_text(strip=True)
                    else:
                        cover_date = cells[2].get_text(strip=True)

                badges = title_cell.find_all('span', class_='badge')
                organisation = ''
                
                for badge in badges:
                    badge_text = badge.get_text(strip=True)
                    if not badge_text.startswith('FP'):
                        organisation = badge_text
                        break

                project_data = {
                    'Titre': titre_base,
                    'Type': document_type,
                    'Document': '',
                    'nom_site': 'greenclimate.fund',
                    'Organisation': organisation,
                    'Lien': project_url,
                    'Description': '',
                    'Cofinancement Total': '',
                    'source': 'GCF',
                    'cover_date': cover_date
                }

                projects.append(project_data)

            except Exception as e:
                logger.error(f"❌ Erreur ligne GCF {i+1}: {e}")
                continue

        return projects

    def enrich_gcf_projects_details(self, projects: List[Dict[str, str]], base_url: str):
        """Enrichit les projets GCF avec les détails de leurs pages individuelles"""
        enriched_projects = []
        
        for i, project in enumerate(projects):
            self.stdout.write(f"🔍 Enrichissement GCF {i+1}/{len(projects)}: {project['Titre'][:30]}...")
            
            detailed_info = self.extract_gcf_project_details(project['Lien'])
            enriched_project = {**project, **detailed_info}
            
            if not enriched_project.get('Description'):
                enriched_project['Description'] = project.get('Description', '')
            
            enriched_projects.append(enriched_project)
            time.sleep(0.5)

        return enriched_projects

    def extract_gcf_project_details(self, project_url: str):
        """Extrait les détails d'un projet GCF depuis sa page dédiée"""
        soup = self.get_page_content(project_url)
        if not soup:
            return {}

        details = {}

        try:
            # Extraction du lien de téléchargement du document PDF
            download_link = soup.find('a', class_='btn-primary', href=True)
            if download_link and download_link.get('href', '').endswith('.pdf'):
                details['Document'] = download_link['href']
            else:
                pdf_links = soup.find_all('a', href=re.compile(r'\.pdf$'))
                if pdf_links:
                    details['Document'] = pdf_links[0]['href']

            # Extraction de la description
            desc_elem = soup.find('div', class_='field-name-body')
            if desc_elem:
                desc_text = desc_elem.find('p')
                if desc_text:
                    details['Description'] = desc_text.get_text(strip=True)
                else:
                    details['Description'] = desc_elem.get_text(strip=True)

            # Extraction des métadonnées
            meta_sections = soup.find_all('div', class_='col-md-6')
            for section in meta_sections:
                label_elem = section.find('span', class_='node-label')
                if not label_elem:
                    continue
                    
                label = label_elem.get_text(strip=True).lower()
                strong_elem = section.find('strong')
                if not strong_elem:
                    continue
                    
                value = strong_elem.get_text(strip=True)
                
                if 'document type' in label or 'type' in label:
                    details['Type'] = value
                elif 'organisation' in label or 'organization' in label:
                    if 'Organisation' not in details:
                        details['Organisation'] = value
                elif any(keyword in label for keyword in ['cofinancement', 'co-financing', 'total funding', 'funding amount', 'amount']):
                    details['Cofinancement Total'] = value
            
            # Recherche alternative pour le cofinancement
            if not details.get('Cofinancement Total'):
                amount_sections = soup.find_all('div', class_=['field-name-field-amount', 'field-name-field-funding'])
                for section in amount_sections:
                    amount_text = section.get_text(strip=True)
                    if amount_text and any(char.isdigit() for char in amount_text):
                        details['Cofinancement Total'] = amount_text
                        break

        except Exception as e:
            logger.error(f"Erreur lors de l'extraction des détails GCF pour {project_url}: {e}")

        return details

    # =============================
    # SAUVEGARDE PAR SOURCE
    # =============================

    def save_projects_by_source(self, projects, source):
        """Sauvegarde les projets dans des fichiers séparés par source"""
        if not projects:
            self.stdout.write("❌ Aucun projet à sauvegarder")
            return

        # Séparer les projets par source
        gef_projects = [p for p in projects if p.get('source', '').upper() == 'GEF']
        gcf_projects = [p for p in projects if p.get('source', '').upper() == 'GCF']
        
        # Sauvegarder GEF si présent
        if gef_projects:
            self.save_single_source_projects(gef_projects, 'GEF')
        
        # Sauvegarder GCF si présent
        if gcf_projects:
            self.save_single_source_projects(gcf_projects, 'GCF')
        
        # Statistiques globales
        self.stdout.write(f"\n📊 STATISTIQUES GLOBALES:")
        self.stdout.write(f"   - Projets GEF sauvegardés: {len(gef_projects)}")
        self.stdout.write(f"   - Projets GCF sauvegardés: {len(gcf_projects)}")
        self.stdout.write(f"   - Total: {len(projects)}")

    def save_single_source_projects(self, projects, source_name):
        """Sauvegarde les projets d'une seule source dans un fichier Excel"""
        if not projects:
            return

        OUTPUT_FILE = Path.cwd() / 'scraped_data' / f'{source_name}_Mauritanie_Projects.xlsx'
        OUTPUT_FILE.parent.mkdir(exist_ok=True)
        
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = f"Projets {source_name} Mauritanie"
        
        # En-têtes standardisés
        headers = [
            "Titre",
            "Type", 
            "Document",
            "nom_site",
            "Organisation",
            "Lien",
            "Description",
            "Cofinancement Total"
        ]
        
        # Couleur spécifique selon la source
        if source_name == 'GEF':
            header_color = "2E7D32"  # Vert foncé pour GEF
        else:  # GCF
            header_color = "1565C0"  # Bleu foncé pour GCF
        
        # Mise en forme des en-têtes
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Largeurs de colonnes optimisées
        column_widths = [60, 30, 70, 20, 25, 60, 80, 25]
        for col, width in enumerate(column_widths, 1):
            worksheet.column_dimensions[worksheet.cell(row=1, column=col).column_letter].width = width
        
        # Ajout des projets
        for project in projects:
            worksheet.append([
                project.get('Titre', ''),
                project.get('Type', ''),
                project.get('Document', ''),
                project.get('nom_site', ''),
                project.get('Organisation', ''),
                project.get('Lien', ''),
                project.get('Description', ''),
                project.get('Cofinancement Total', '')
            ])
        
        # Formatage avancé des cellules de données
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                # Bordures légères
                from openpyxl.styles import Border, Side
                thin_border = Border(
                    left=Side(style='thin', color='CCCCCC'),
                    right=Side(style='thin', color='CCCCCC'),
                    top=Side(style='thin', color='CCCCCC'),
                    bottom=Side(style='thin', color='CCCCCC')
                )
                cell.border = thin_border
        
        # Ajustement automatique de la hauteur des lignes
        for row_num in range(2, worksheet.max_row + 1):
            worksheet.row_dimensions[row_num].height = 50
        
        # Figer la première ligne
        worksheet.freeze_panes = "A2"
        
        # Sauvegarde
        workbook.save(OUTPUT_FILE)
        
        self.stdout.write(f"✅ {source_name}: {len(projects)} projets sauvegardés dans {OUTPUT_FILE}")

    def save_consolidated_projects(self, projects, source):
        """Sauvegarde consolidée des projets GEF et GCF (optionnelle)"""
        if not projects:
            self.stdout.write("❌ Aucun projet à sauvegarder")
            return

        OUTPUT_FILE = Path.cwd() / 'scraped_data' / f'Mauritanie_Projects_CONSOLIDATED.xlsx'
        OUTPUT_FILE.parent.mkdir(exist_ok=True)
        
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Tous Projets Mauritanie"
        
        # En-têtes avec colonne Source
        headers = [
            "Titre",
            "Type", 
            "Document",
            "nom_site",
            "Organisation",
            "Lien",
            "Description",
            "Cofinancement Total",
            "Source"
        ]
        
        # Mise en forme des en-têtes
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="6A1B9A", end_color="6A1B9A", fill_type="solid")  # Violet pour consolidé
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Largeurs de colonnes
        column_widths = [60, 30, 70, 20, 25, 60, 80, 25, 15]
        for col, width in enumerate(column_widths, 1):
            worksheet.column_dimensions[worksheet.cell(row=1, column=col).column_letter].width = width
        
        # Ajout des projets avec formatage conditionnel par source
        for project in projects:
            row_data = [
                project.get('Titre', ''),
                project.get('Type', ''),
                project.get('Document', ''),
                project.get('nom_site', ''),
                project.get('Organisation', ''),
                project.get('Lien', ''),
                project.get('Description', ''),
                project.get('Cofinancement Total', ''),
                project.get('source', '').upper()
            ]
            worksheet.append(row_data)
            
            # Coloration différente selon la source
            current_row = worksheet.max_row
            source_color = "E8F5E8" if project.get('source', '').upper() == 'GEF' else "E3F2FD"  # Vert clair ou bleu clair
            
            for col in range(1, len(headers) + 1):
                cell = worksheet.cell(row=current_row, column=col)
                cell.fill = PatternFill(start_color=source_color, end_color=source_color, fill_type="solid")
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        
        # Sauvegarde
        workbook.save(OUTPUT_FILE)
        
        # Statistiques
        gef_count = len([p for p in projects if p.get('source', '').upper() == 'GEF'])
        gcf_count = len([p for p in projects if p.get('source', '').upper() == 'GCF'])
        
        self.stdout.write(f"✅ CONSOLIDÉ: {len(projects)} projets sauvegardés dans {OUTPUT_FILE}")
        self.stdout.write(f"   - GEF: {gef_count} | GCF: {gcf_count}")


# =============================
# FONCTIONS UTILITAIRES
# =============================

def run_integrated_scraper(source='both', max_pages=50, headless=False, max_details=None):
    """Fonction utilitaire pour lancer le scraper intégré GEF-GCF"""
    command = Command()
    
    class MockOptions:
        def get(self, key, default=None):
            options_dict = {
                'source': source,
                'max_pages': max_pages,
                'headless': headless,
                'max_details': max_details
            }
            return options_dict.get(key, default)
    
    options = MockOptions()
    command.handle(options=options)

def test_connections():
    """Test de connexion aux sites GEF et GCF"""
    print("🔍 Test de connexion aux sites...")
    
    # Test GEF
    try:
        response = requests.get("https://www.thegef.org/projects-operations/database", timeout=10)
        if response.status_code == 200:
            print("✅ GEF: Connexion réussie!")
            gef_ok = True
        else:
            print(f"❌ GEF: Erreur {response.status_code}")
            gef_ok = False
    except Exception as e:
        print(f"❌ GEF: Impossible de se connecter - {e}")
        gef_ok = False
    
    # Test GCF
    try:
        response = requests.get("https://www.greenclimate.fund/countries/mauritania", timeout=10)
        if response.status_code == 200:
            print("✅ GCF: Connexion réussie!")
            gcf_ok = True
        else:
            print(f"❌ GCF: Erreur {response.status_code}")
            gcf_ok = False
    except Exception as e:
        print(f"❌ GCF: Impossible de se connecter - {e}")
        gcf_ok = False
    
    return gef_ok, gcf_ok

def main():
    """Fonction principale pour exécuter le scraper intégré"""
    print("🚀 Démarrage du Scraper Intégré GEF-GCF Mauritanie")
    print("="*60)
    
    # Test de connexion
    gef_ok, gcf_ok = test_connections()
    
    if not gef_ok and not gcf_ok:
        print("🚫 Impossible de continuer sans connexion aux sites")
        return
    
    # Déterminer quelle source utiliser
    if gef_ok and gcf_ok:
        source = 'both'
        print("✅ Les deux sources sont disponibles")
    elif gef_ok:
        source = 'gef'
        print("⚠️ Seul GEF est disponible")
    else:
        source = 'gcf'
        print("⚠️ Seul GCF est disponible")
    
    try:
        # Configuration de l'extraction
        run_integrated_scraper(
            source=source,           # Source déterminée automatiquement
            max_pages=20,           # Pages GEF à scraper
            headless=False,         # Mode visible pour debugging
            max_details=None        # Pas de limite pour GCF
        )
        
        print("\n" + "="*60)
        print("🎉 EXTRACTION TERMINÉE AVEC SUCCÈS!")
        print("="*60)
        
    except Exception as e:
        print(f"\n❌ Erreur lors de l'exécution: {e}")
        logger.error(f"Erreur principale: {e}")

# =============================
# CLASSE POUR USAGE STANDALONE
# =============================

class StandaloneIntegratedScraper:
    """Version standalone du scraper intégré (sans Django)"""
    
    def __init__(self):
        self.gef_scraper = None
        self.gcf_session = requests.Session()
        self.gcf_session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        })
    
    def run_scraping(self, source='both', max_pages=20, headless=False, max_details=None):
        """Lance le scraping sans Django"""
        print(f"🚀 Scraping {source.upper()} - Mauritanie")
        
        all_projects = []
        
        if source in ['gef', 'both']:
            print("\n🌍 === SCRAPING GEF ===")
            # Pour GEF, utiliser la logique Selenium du Command
            command = Command()
            try:
                command.setup_driver(headless)
                gef_projects = command.scrape_gef_mauritania_projects(max_pages)
                all_projects.extend(gef_projects)
                command.cleanup_driver()
            except Exception as e:
                print(f"❌ Erreur GEF: {e}")
        
        if source in ['gcf', 'both']:
            print("\n🌱 === SCRAPING GCF ===")
            # Pour GCF, utiliser la logique requests
            command = Command()
            try:
                gcf_projects = command.scrape_gcf_mauritania_projects(max_details)
                all_projects.extend(gcf_projects)
            except Exception as e:
                print(f"❌ Erreur GCF: {e}")
        
        # Sauvegarde séparée
        if all_projects:
            command = Command()
            command.save_projects_by_source(all_projects, source)
        
        return all_projects

# =============================
# EXEMPLES D'UTILISATION
# =============================

if __name__ == "__main__":
    # Option 1: Utilisation avec Django (tâche planifiée)
    # python manage.py scrape_mauritania --source=both --max-pages=20
    
    # Option 2: Utilisation standalone
    main()
    
    # Option 3: Utilisation programmatique
    # scraper = StandaloneIntegratedScraper()
    # projects = scraper.run_scraping(source='gcf', max_details=10)
    
    # Option 4: GEF seulement
    # run_integrated_scraper(source='gef', max_pages=10, headless=True)
    
    # Option 6: Sauvegarder aussi un fichier consolidé (optionnel)
    # command = Command()
    # command.save_consolidated_projects(all_projects, 'both')